# Save as "extract_aqi.py"

import os
from pathlib import Path
import PyPDF2
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class AQIDataExtractor:
    def __init__(self, pdf_dir):
        self.pdf_dir = Path(pdf_dir)
        if not self.pdf_dir.exists():
            self.pdf_dir.mkdir(parents=True, exist_ok=True)
        
        self.aqi_colors = {
            'Good': '00FF00',
            'Satisfactory': '90EE90',
            'Moderate': 'FFFF00',
            'Poor': 'FFA500',
            'Very Poor': 'FF0000',
            'Severe': '800080'
        }

    def find_data_in_text(self, text, city):
        """Extract AQI data from text using pattern matching"""
        try:
            lines = text.split('\n')
            line_count = len(lines)
            
            # Look for lines containing the city
            for i, line in enumerate(lines):
                if city.lower() in line.lower():
                    # Get current line
                    city_line = line.strip()
                    
                    # Get a few lines after for context (up to 3 lines)
                    after_lines = []
                    for j in range(1, 4):
                        if i + j < line_count:
                            after_lines.append(lines[i + j].strip())
                    
                    # Combine lines for searching
                    search_text = ' '.join([city_line] + after_lines)
                    print(f"\nAnalyzing text: {search_text}")
                    
                    data = {}
                    
                    # Extract Air Quality
                    for category in self.aqi_colors:
                        if category in search_text:
                            data['Air_Quality'] = category
                            data['Color'] = self.aqi_colors[category]
                            print(f"Found Air Quality: {category}")
                            break
                    
                    # Extract Index Value (look for 2-3 digit number)
                    index_matches = re.findall(r'\b(\d{2,3})\b', search_text)
                    for value in index_matches:
                        if 50 <= int(value) <= 500:
                            data['Index_Value'] = value
                            print(f"Found Index Value: {value}")
                            break

                    # Extract Pollutants with improved pattern matching
                    pollutant_text = search_text.replace(' ', '')  # Remove spaces for better matching
                    pollutants = []

                    # Look for specific pollutant patterns
                    if 'PM2.5' in pollutant_text:
                        pollutants.append('PM₂.₅')
                    if 'PM10' in pollutant_text:
                        pollutants.append('PM₁₀')
                    if 'O3' in pollutant_text:
                        pollutants.append('O₃')
                    if 'NO2' in pollutant_text:
                        pollutants.append('NO₂')
                    if 'SO2' in pollutant_text:
                        pollutants.append('SO₂')
                    if 'CO' in pollutant_text:
                        pollutants.append('CO')

                    # Combine found pollutants
                    if pollutants:
                        data['Prominent_Pollutant'] = ', '.join(pollutants)
                        print(f"Found Pollutants: {data['Prominent_Pollutant']}")

                    if 'Air_Quality' in data and 'Index_Value' in data:
                        print(f"Found complete data: {data}")
                        return data
            
            return None
        except Exception as e:
            print(f"Error extracting data: {e}")
            return None

    def process_pdf(self, pdf_path, city):
        """Process a single PDF file"""
        try:
            if not pdf_path.exists():
                print(f"File not found: {pdf_path}")
                return None

            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    text = page.extract_text()
                    if city.lower() in text.lower():
                        data = self.find_data_in_text(text, city)
                        if data:
                            return data
            return None
        except Exception as e:
            print(f"Error processing PDF: {e}")
            return None

    def process_date_range(self, city, start_date, end_date):
        """Process PDFs for the given date range"""
        data_list = []
        current_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        
        while current_date <= end_dt:
            date_str = current_date.strftime("%Y%m%d")
            pdf_path = self.pdf_dir / f"AQI_Bulletin_{date_str}.pdf"
            
            print(f"\nProcessing date: {current_date.strftime('%Y-%m-%d')}")
            
            if pdf_path.exists():
                data = self.process_pdf(pdf_path, city)
                if data:
                    data['Date'] = current_date.strftime('%Y-%m-%d')
                    data_list.append(data)
                    print(f"Successfully extracted data for {current_date.strftime('%Y-%m-%d')}")
            else:
                print(f"PDF not found: {pdf_path}")
            
            current_date += timedelta(days=1)
        
        return data_list

    def create_excel(self, data_list, city, start_date, end_date):
        """Create Excel file with extracted data"""
        if not data_list:
            print("No data to write to Excel")
            return False

        filename = f"{city}_AQI_{start_date}_to_{end_date}.xlsx"
        wb = Workbook()
        ws = wb.active

        # Title
        title = f"{city} AQI Information from {start_date} to {end_date}"
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells('A1:D1')
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # Add blank row
        ws.append([])

        # Headers
        headers = ['Date', 'Air Quality', 'Index Value', 'Prominent Pollutant']
        ws.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True)

        # Add data
        for data in data_list:
            row = [
                data['Date'],
                data['Air_Quality'],
                data['Index_Value'],
                data.get('Prominent_Pollutant', '')
            ]
            ws.append(row)
            
            # Color the Air Quality cell
            last_row = ws.max_row
            cell = ws.cell(row=last_row, column=2)
            cell.fill = PatternFill(start_color=data['Color'], 
                                  end_color=data['Color'], 
                                  fill_type='solid')
            
            # Set text color based on background
            rgb = int(data['Color'], 16)
            brightness = (((rgb >> 16) & 255) * 299 + 
                        ((rgb >> 8) & 255) * 587 + 
                        (rgb & 255) * 114) / 1000
            cell.font = Font(color='000000' if brightness > 128 else 'FFFFFF')

        # Set column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.row > 1:  # Skip title row
                    try:
                        max_length = max(max_length, len(str(cell.value or '')))
                    except:
                        pass
            if max_length > 0:
                ws.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width = max_length + 2

        try:
            wb.save(filename)
            print(f"\nExcel file created: {filename}")
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False

def main():
    try:
        print("\nAQI Data Extractor")
        print("=================\n")
        
        # Get user input
        pdf_dir = input("Enter directory containing PDF files (default: aqi_bulletins): ").strip() or "aqi_bulletins"
        city = input("Enter city name: ").strip()
        start_date = input("Enter start date (YYYY-MM-DD): ").strip()
        end_date = input("Enter end date (YYYY-MM-DD): ").strip()
        
        if not city:
            print("Error: City name is required")
            return
        
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            print("Error: Invalid date format. Use YYYY-MM-DD")
            return
        
        # Process files
        extractor = AQIDataExtractor(pdf_dir)
        data = extractor.process_date_range(city, start_date, end_date)
        
        if data:
            extractor.create_excel(data, city, start_date, end_date)
        else:
            print("\nNo data found for the specified city and date range")
        
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
    except Exception as e:
        print(f"\nAn error occurred: {e}")

if __name__ == "__main__":
    main()
